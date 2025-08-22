from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from datetime import datetime, date
import os
import sqlite3
from flask_socketio import SocketIO

app = Flask(__name__)
socketio = SocketIO(app)

generated_file = "generated/updated_trim.xlsx"
live_users = 0


# ===================== DATABASE HELPERS =====================
def init_db():
    conn = sqlite3.connect("stats.db")
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS sheets (id INTEGER PRIMARY KEY, timestamp DATE)")
    conn.commit()
    conn.close()

def log_sheet_generated():
    conn = sqlite3.connect("stats.db")
    c = conn.cursor()
    c.execute("INSERT INTO sheets (timestamp) VALUES (?)", (date.today().isoformat(),))
    conn.commit()
    conn.close()

def get_today_count():
    conn = sqlite3.connect("stats.db")
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM sheets WHERE timestamp = ?", (date.today().isoformat(),))
    count = c.fetchone()[0]
    conn.close()
    return count


# ===================== TRIM SHEET GENERATORS =====================
def generate_trim_sheet_152(regn, pilot_weight, pax_weight, fuel_left, fuel_right):
    wb = load_workbook('master_trim_152.xlsx')
    ws = wb.active

    ws["B1"] = datetime.now().strftime("%d/%m/%Y")
    ws["E2"] = regn

    if regn == "IAU":
        c5, e5 = 1173.96, 31.47
    elif regn == "NNN":
        c5, e5 = 1219, 31.6
    elif regn == "PSS":
        c5, e5 = 1201, 31.09
    else:
        c5, e5 = 0, 0

    ws["C5"] = round(c5, 2)
    ws["E5"] = round(e5, 2)
    ws["G5"] = round(c5 * e5, 2)

    ws["C6"] = round(pilot_weight, 2)
    ws["C7"] = round(pax_weight, 2)

    e6 = ws["E6"].value or 0
    e7 = ws["E7"].value or 0

    ws["G6"] = round(pilot_weight * e6, 2)
    ws["G7"] = round(pax_weight * e7, 2)

    ws["B18"] = round(c5 + pilot_weight + pax_weight, 2)
    ws["E11"] = round(ws["G5"].value + ws["G6"].value + ws["G7"].value, 2)

    ws["B13"] = round(fuel_left, 2)
    ws["B14"] = round(fuel_right, 2)
    ws["B15"] = round(fuel_left + fuel_right, 2)

    ws["C13"] = round(fuel_left * 1.58, 2)
    ws["C14"] = round(fuel_right * 1.58, 2)
    ws["C15"] = round(ws["C13"].value + ws["C14"].value, 2)

    e15 = ws["E15"].value or 0
    ws["G15"] = round(ws["C15"].value * e15, 2)

    ws["E12"] = round(ws["G15"].value, 2)
    ws["E13"] = round(ws["E11"].value + ws["E12"].value, 2)

    ws["C20"] = round(ws["C5"].value + ws["C6"].value + ws["C7"].value + ws["C13"].value + ws["C14"].value, 2)
    ws["G20"] = "Y" if ws["C20"].value < 1670 else "N"

    if ws["C20"].value and ws["E13"].value:
        ws["C21"] = round(ws["E13"].value / ws["C20"].value, 2)

    cg_val = ws["C21"].value
    ws["G21"] = "Y" if cg_val is not None and 31 <= cg_val <= 36.5 else "N"

    os.makedirs("generated", exist_ok=True)
    wb.save(generated_file)

    data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=24, max_col=7)]
    return data


def generate_trim_sheet_172(regn, pilot_weight, pax_weight, fuel_left, fuel_right):
    wb = load_workbook('master_trim_172.xlsx')
    ws = wb.active

    ws["B1"] = datetime.now().strftime("%d/%m/%Y")

    if regn == "AGH":
        ws["D2"] = "VT-AGH"
        c5, e5 = 1697, 39.29
    elif regn == "PFA":
        ws["D2"] = "VT-PFA"
        c5, e5 = 1701, 38.79
    else:
        c5, e5 = 0, 0

    ws["C5"] = round(c5, 2)
    ws["E5"] = round(e5, 2)
    ws["G5"] = round(c5 * e5, 2)

    ws["C6"] = round(pilot_weight, 2)
    ws["C7"] = round(pax_weight, 2)

    e6 = ws["E6"].value or 0
    e7 = ws["E7"].value or 0
    ws["G6"] = round(pilot_weight * e6, 2)
    ws["G7"] = round(pax_weight * e7, 2)

    ws["B15"] = round(fuel_left, 2)
    ws["B16"] = round(fuel_right, 2)

    ws["C15"] = round(fuel_left * 1.58, 2)
    ws["C16"] = round(fuel_right * 1.58, 2)

    ws["B17"] = round(fuel_left + fuel_right, 2)
    ws["C17"] = round(ws["C15"].value + ws["C16"].value, 2)

    e17 = ws["E17"].value or 0
    ws["G17"] = round(ws["C17"].value * e17, 2)

    ws["E13"] = round(ws["G5"].value + ws["G6"].value + ws["G7"].value, 2)
    ws["E14"] = round(ws["G17"].value, 2)
    ws["E15"] = round(ws["E13"].value + ws["E14"].value, 2)

    ws["B19"] = round(c5 + pilot_weight + pax_weight, 2)
    ws["C21"] = round(c5 + pilot_weight + pax_weight + ws["C17"].value, 2)

    ws["G21"] = "Y" if ws["C21"].value <= 2550 else "N"

    if ws["C21"].value and ws["E15"].value:
        ws["C22"] = round(ws["E15"].value / ws["C21"].value, 2)

    cg_val = ws["C22"].value
    ws["G22"] = "Y" if cg_val is not None and 35 <= cg_val <= 47.4 else "N"

    os.makedirs("generated", exist_ok=True)
    wb.save(generated_file)

    data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=24, max_col=7)]
    return data


# ===================== ROUTES =====================
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        regn = request.form['regn']
        pilot_weight = float(request.form['pilot_weight_lbs'])
        pax_weight = float(request.form['pax_weight_lbs'])
        fuel_left = float(request.form['fuel_left'])
        fuel_right = float(request.form['fuel_right'])

        if regn in ["IAU", "NNN", "PSS"]:
            data = generate_trim_sheet_152(regn, pilot_weight, pax_weight, fuel_left, fuel_right)
        elif regn in ["AGH", "PFA"]:
            data = generate_trim_sheet_172(regn, pilot_weight, pax_weight, fuel_left, fuel_right)
        else:
            return "Invalid Aircraft Registration", 400

        # Log the generation
        log_sheet_generated()

        return render_template("table.html", data=data, excel_available=True)

    return render_template("index.html")


@app.route('/download_excel')
def download_excel():
    return send_file(generated_file, as_attachment=True)


@app.route('/today_count')
def today_count():
    return {"today_count": get_today_count()}


# ===================== SOCKET EVENTS =====================
@socketio.on('connect')
def handle_connect():
    global live_users
    live_users += 1
    socketio.emit('update_user_count', live_users, broadcast=True)

@socketio.on('disconnect')
def handle_disconnect():
    global live_users
    live_users -= 1
    socketio.emit('update_user_count', live_users, broadcast=True)


# ===================== MAIN =====================
if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    socketio.run(app, host="0.0.0.0", port=port, debug=True)
